"""
export_excel.py - build a formatted .xlsx from the current screen view.

The UI writes the rows it currently shows (filters + pins + sort already
applied) to data/export_screen.csv, and the trades to data/trades.csv,
then calls this script. We produce a polished workbook with:
  - styled header, frozen top row, autofilter
  - number formats, auto-fit columns
  - conditional formatting (deep-drop rows, P/L coloring)

Usage:
    python export_excel.py <output_path.xlsx>

Reads:
    data/export_screen.csv   (written by the UI just before calling this)
    data/trades.csv          (the paper-trading log)
"""

import os
import sys
import csv
import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule
except ImportError:
    sys.stderr.write("openpyxl not installed. Run: pip install openpyxl\n")
    sys.exit(2)

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, "data")
SCREEN_CSV = os.path.join(DATA_DIR, "export_screen.csv")
TRADES_CSV = os.path.join(DATA_DIR, "trades.csv")

NAVY = "1F3864"
MID = "2F5496"
WHITE = "FFFFFF"
ALT = "F2F5FA"
REDF = "FCE4E4"
GREENF = "E2EFDA"

hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
title_font = Font(name="Calibri", bold=True, color=WHITE, size=14)
cell_font = Font(name="Calibri", size=10)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill(c):
    return PatternFill("solid", start_color=c)


def read_csv(path):
    if not os.path.exists(path):
        return [], []
    with open(path, newline="", encoding="utf-8") as f:
        r = csv.reader(f)
        rows = list(r)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def autosize(ws, header, data, max_w=42):
    for ci, h in enumerate(header, 1):
        longest = len(str(h))
        for row in data:
            if ci - 1 < len(row):
                longest = max(longest, len(str(row[ci - 1])))
        ws.column_dimensions[get_column_letter(ci)].width = min(longest + 3, max_w)


def num_fmt_for(colname):
    c = colname.lower()
    if "price" in c or "high" in c or "value" in c or "investment" in c or "p/l $" in c:
        return '#,##0.00'
    if "%" in c or "pct" in c or "below" in c or "chg" in c or "p/l %" in c:
        return '0.0"%"'
    if "cap" in c or "vol" in c:
        return '#,##0.0'
    return None


def build_sheet(wb, title, header, data, first=True):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.sheet_view.showGridLines = False

    ncol = max(1, len(header))
    last_col = get_column_letter(ncol)

    # Title band
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"{title.upper()}   -   exported {datetime.datetime.now():%Y-%m-%d %H:%M}"
    ws["A1"].font = title_font
    ws["A1"].fill = fill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, ncol + 1):
        ws.cell(1, c).fill = fill(NAVY)
    ws.row_dimensions[1].height = 24

    # Header row (row 3, leaving row 2 as a spacer)
    hdr_row = 3
    for ci, h in enumerate(header, 1):
        cell = ws.cell(hdr_row, ci, h)
        cell.font = hdr_font
        cell.fill = fill(MID)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[hdr_row].height = 26

    # Data
    for ri, row in enumerate(data):
        excel_row = hdr_row + 1 + ri
        for ci in range(ncol):
            val = row[ci] if ci < len(row) else ""
            # numeric coercion where sensible
            num = None
            if val not in ("", "-", None):
                try:
                    num = float(str(val).replace(",", "").replace("%", "").replace("$", ""))
                except ValueError:
                    num = None
            cell = ws.cell(excel_row, ci + 1, num if num is not None else val)
            cell.font = cell_font
            cell.border = border
            fmt = num_fmt_for(header[ci]) if ci < len(header) else None
            if num is not None and fmt:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
            if ri % 2:
                if cell.fill.start_color.rgb in (None, "00000000"):
                    cell.fill = fill(ALT)

    n_data = len(data)
    last_data_row = hdr_row + n_data
    if n_data > 0:
        ws.freeze_panes = ws.cell(hdr_row + 1, 1)
        ws.auto_filter.ref = f"A{hdr_row}:{last_col}{last_data_row}"

        # Conditional formatting: highlight "% Below" > 70 (deep drop) if present
        for ci, h in enumerate(header):
            if "below" in h.lower():
                col = get_column_letter(ci + 1)
                rng = f"{col}{hdr_row+1}:{col}{last_data_row}"
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(formula=[f"{col}{hdr_row+1}>70"],
                                stopIfTrue=False, fill=fill(REDF)))
            if "p/l $" in h.lower():
                col = get_column_letter(ci + 1)
                rng = f"{col}{hdr_row+1}:{col}{last_data_row}"
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(formula=[f"{col}{hdr_row+1}>0"],
                                stopIfTrue=False, fill=fill(GREENF)))
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(formula=[f"{col}{hdr_row+1}<0"],
                                stopIfTrue=False, fill=fill(REDF)))

    autosize(ws, header, data)
    return ws


def main():
    if len(sys.argv) < 2:
        sys.stderr.write("Usage: python export_excel.py <output.xlsx>\n")
        sys.exit(1)
    out_path = sys.argv[1]

    s_hdr, s_data = read_csv(SCREEN_CSV)
    t_hdr, t_data = read_csv(TRADES_CSV)

    if not s_hdr:
        sys.stderr.write("No screen data to export.\n")
        sys.exit(1)

    wb = Workbook()
    build_sheet(wb, "Screen", s_hdr, s_data, first=True)
    if t_hdr:
        build_sheet(wb, "Test Trades", t_hdr, t_data, first=False)

    # atomic-ish write
    tmp = out_path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, out_path)
    sys.stdout.write(f"Exported {len(s_data)} screen rows -> {out_path}\n")


if __name__ == "__main__":
    main()
